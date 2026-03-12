# -*- coding: utf-8 -*-
"""엑셀 파일 내 모든 거래처 이메일을 sypark.dpk@gmail.com 으로 변경합니다."""
import shutil
from pathlib import Path

NEW_EMAIL = "sypark.dpk@gmail.com"
EXCEL_NAME = "domino_inventory_training.xlsx"


def update_with_openpyxl(path: Path) -> bool:
    """openpyxl로 기존 파일을 열어 이메일 셀만 교체."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path)
        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    s = str(cell.value).strip()
                    if "@" in s and (".com" in s or ".co.kr" in s or "naver" in s or "gmail" in s):
                        cell.value = NEW_EMAIL
        wb.save(path)
        return True
    except Exception as e:
        print("openpyxl 로 수정 실패:", e)
        return False


def create_new_workbook(path: Path) -> None:
    """동일 구조의 새 엑셀을 생성하고 모든 이메일을 NEW_EMAIL로 채움."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Guide
    ws_guide = wb.create_sheet("Guide", 0)
    guide_data = [
        ["도미노피자 재고·발주 자동화 패키지"],
        [],
        ["패키지 사용 안내"],
        [],
        ["실습 목표", "재고를 확인하고, 재고 부족 품목을 찾고, 담당자에게 알림을 보여준 뒤 발주서와 이메일 초안을 만드는 MVP를 구현합니다."],
        ["재고 부족 기준", "현재재고 < 안전재고"],
        ["발주 권장 수량", "현재재고가 안전재고보다 적으면 MAX(MOQ, 안전재고-현재재고)"],
        [], [],
        ["재료 10개"],
        [1, "도우볼", None, "주의"],
        [2, "토마토소스"],
        [3, "모짜렐라치즈"],
        [4, "페퍼로니"],
        [5, "베이컨"],
        [6, "양파"],
        [7, "피망"],
        [8, "양송이버섯"],
        [9, "블랙올리브"],
        [10, "스위트콘"],
    ]
    for r, row in enumerate(guide_data, 1):
        for c, val in enumerate(row if isinstance(row, (list, tuple)) else [row], 1):
            ws_guide.cell(row=r, column=c, value=val)

    # Suppliers (이메일 모두 NEW_EMAIL)
    ws_sup = wb.create_sheet("Suppliers", 1)
    sup_headers = ["거래처명", "담당자", "이메일", "리드타임(일)", "품목군"]
    sup_data = [
        ["도미노푸드서플라이", "박지훈", NEW_EMAIL, 2, "도우/소스/치즈"],
        ["프레시미트코리아", "김현우", NEW_EMAIL, 3, "페퍼로니/베이컨"],
        ["그린베지유통", "이수진", NEW_EMAIL, 1, "양파/피망/버섯"],
        ["토핑솔루션", "정민아", NEW_EMAIL, 2, "올리브/콘"],
    ]
    for c, h in enumerate(sup_headers, 1):
        ws_sup.cell(row=1, column=c, value=h)
    for r, row in enumerate(sup_data, 2):
        for c, val in enumerate(row, 1):
            ws_sup.cell(row=r, column=c, value=val)

    # Inventory (거래처이메일 모두 NEW_EMAIL)
    ws_inv = wb.create_sheet("Inventory", 2)
    inv_headers = [
        "품목코드", "재료명", "규격", "단위", "현재재고", "안전재고", "MOQ",
        "거래처", "알림담당자", "거래처이메일", "리드타임(일)", "부족수량", "발주권장수량", "상태", "담당자알림메시지"
    ]
    inv_data = [
        ["ING001", "도우볼", "220g", "개", 120, 180, 100, "도미노푸드서플라이", "점포 운영매니저", NEW_EMAIL, 2, 60, 100, "발주 필요", "도우볼 재고 부족 - 현재 120개, 안전재고 180개, 권장발주 100개"],
        ["ING002", "토마토소스", "3kg", "팩", 32, 20, 10, "도미노푸드서플라이", "점포 운영매니저", NEW_EMAIL, 2, 0, 0, "정상", None],
        ["ING003", "모짜렐라치즈", "2kg", "봉", 12, 18, 10, "도미노푸드서플라이", "점포 운영매니저", NEW_EMAIL, 2, 6, 10, "발주 필요", "모짜렐라치즈 재고 부족 - 현재 12봉, 안전재고 18봉, 권장발주 10봉"],
        ["ING004", "페퍼로니", "1kg", "팩", 15, 12, 8, "프레시미트코리아", "점포 운영매니저", NEW_EMAIL, 3, 0, 0, "정상", None],
        ["ING005", "베이컨", "1kg", "팩", 6, 10, 10, "프레시미트코리아", "점포 운영매니저", NEW_EMAIL, 3, 4, 10, "발주 필요", "베이컨 재고 부족 - 현재 6팩, 안전재고 10팩, 권장발주 10팩"],
        ["ING006", "양파", "5kg", "봉", 14, 12, 6, "그린베지유통", "점포 운영매니저", NEW_EMAIL, 1, 0, 0, "정상", None],
        ["ING007", "피망", "5kg", "봉", 5, 8, 5, "그린베지유통", "점포 운영매니저", NEW_EMAIL, 1, 3, 5, "발주 필요", "피망 재고 부족 - 현재 5봉, 안전재고 8봉, 권장발주 5봉"],
        ["ING008", "양송이버섯", "2.5kg", "캔", 3, 6, 6, "그린베지유통", "점포 운영매니저", NEW_EMAIL, 1, 3, 6, "발주 필요", "양송이버섯 재고 부족 - 현재 3캔, 안전재고 6캔, 권장발주 6캔"],
        ["ING009", "블랙올리브", "3kg", "캔", 9, 8, 4, "토핑솔루션", "점포 운영매니저", NEW_EMAIL, 2, 0, 0, "정상", None],
        ["ING010", "스위트콘", "3kg", "캔", 4, 5, 6, "토핑솔루션", "점포 운영매니저", NEW_EMAIL, 2, 1, 6, "발주 필요", "스위트콘 재고 부족 - 현재 4캔, 안전재고 5캔, 권장발주 6캔"],
    ]
    for c, h in enumerate(inv_headers, 1):
        ws_inv.cell(row=1, column=c, value=h)
    for r, row in enumerate(inv_data, 2):
        for c, val in enumerate(row, 1):
            ws_inv.cell(row=r, column=c, value=val)

    # OrderSummary (담당자 이메일 모두 NEW_EMAIL)
    ws_ord = wb.create_sheet("OrderSummary", 3)
    ord_data = [
        ["도미노피자 발주 요약"],
        [],
        ["총 품목 수", 10],
        ["발주 필요 품목 수", 6],
        ["전체 권장 발주 수량", 137],
        ["오늘 상태", "담당자 확인 필요"],
        [], [],
        ["거래처별 발주 필요 현황"],
        [],
        ["거래처명", "발주 필요 품목 수", "총 권장 발주 수량", "담당자 이메일", "리드타임(일)"],
        ["도미노푸드서플라이", 2, 110, NEW_EMAIL, 2],
        ["프레시미트코리아", 1, 10, NEW_EMAIL, 3],
        ["그린베지유통", 2, 11, NEW_EMAIL, 1],
        ["토핑솔루션", 1, 6, NEW_EMAIL, 2],
    ]
    for r, row in enumerate(ord_data, 1):
        row_list = row if isinstance(row, (list, tuple)) else [row]
        for c, val in enumerate(row_list, 1):
            ws_ord.cell(row=r, column=c, value=val)

    # EmailTemplate
    ws_em = wb.create_sheet("EmailTemplate", 4)
    em_data = [
        ["발주 이메일 템플릿"],
        [],
        ["제목 템플릿", "[발주요청] {{STORE_NAME}} / {{SUPPLIER_NAME}} / {{ORDER_DATE}}"],
        [],
        ["본문 템플릿", "안녕하세요 {{SUPPLIER_NAME}} 담당자님.\n\n도미노피자 {{STORE_NAME}}입니다.\n아래 품목에 대해 발주 요청드립니다.\n\n{{ITEM_LIST}}\n\n첨부한 발주서 확인 부탁드립니다.\n감사합니다.\n{{INTERNAL_OWNER}}"],
    ]
    for r, row in enumerate(em_data, 1):
        row_list = row if isinstance(row, (list, tuple)) else [row]
        for c, val in enumerate(row_list, 1):
            ws_em.cell(row=r, column=c, value=val)

    wb.save(path)
    print("새 엑셀 파일을 생성했습니다:", path)


def main():
    base = Path(__file__).resolve().parent
    path = base / EXCEL_NAME

    # 1) 기존 파일이 있으면 openpyxl로 수정 시도
    if path.exists():
        backup = base / (EXCEL_NAME.replace(".xlsx", "_backup.xlsx"))
        try:
            shutil.copy2(path, backup)
            print("기존 파일 백업:", backup)
        except Exception as e:
            print("백업 실패 (무시하고 진행):", e)
        if update_with_openpyxl(path):
            print("기존 파일 수정 완료. 모든 거래처 이메일 →", NEW_EMAIL)
            return
    else:
        path.parent.mkdir(parents=True, exist_ok=True)

    # 2) 실패 시 또는 파일 없음 시 동일 구조로 새 파일 생성
    print("동일 구조로 새 엑셀을 생성합니다 (모든 이메일:", NEW_EMAIL, ")")
    create_new_workbook(path)
    print("완료.")


if __name__ == "__main__":
    main()
